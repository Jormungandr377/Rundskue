"""Data export endpoints for transactions and reports (CSV, Excel, PDF)."""
from datetime import date, datetime, timezone
from typing import Optional
from io import BytesIO
import csv

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from slowapi import Limiter
from slowapi.util import get_remote_address
from sqlalchemy.orm import Session, joinedload

from ..database import get_db
from ..models import Transaction, Account, Category, Profile, NetWorthSnapshot
from ..dependencies import get_current_active_user
from ..services import audit

router = APIRouter(tags=["Export"])
limiter = Limiter(key_func=get_remote_address)


def get_user_profile(db: Session, user) -> Profile:
    """Get the primary profile for the current user."""
    profile = db.query(Profile).filter(
        Profile.user_id == user.id,
        Profile.is_primary == True
    ).first()
    if not profile:
        raise HTTPException(status_code=404, detail="No primary profile found")
    return profile


@router.get("/transactions/csv")
@limiter.limit("10/minute")
async def export_transactions_csv(
    request: Request,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user=Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Export transactions as CSV file."""
    profile = get_user_profile(db, current_user)

    # Get accounts for this profile
    account_ids = [a.id for a in db.query(Account).filter(Account.profile_id == profile.id).all()]

    query = db.query(Transaction).options(
        joinedload(Transaction.account),
        joinedload(Transaction.category)
    ).filter(Transaction.account_id.in_(account_ids))

    if start_date:
        query = query.filter(Transaction.date >= start_date)
    if end_date:
        query = query.filter(Transaction.date <= end_date)

    transactions = query.order_by(Transaction.date.desc()).limit(50000).all()

    # Build CSV
    output = BytesIO()
    import io
    text_output = io.TextIOWrapper(output, encoding='utf-8', newline='')
    writer = csv.writer(text_output)

    # Header
    writer.writerow([
        "Date", "Description", "Account", "Category", "Amount",
        "Type", "Excluded", "Transfer", "Notes"
    ])

    for txn in transactions:
        writer.writerow([
            str(txn.date),
            txn.custom_name or txn.merchant_name or txn.name,
            txn.account.display_name or txn.account.name if txn.account else "",
            txn.category.name if txn.category else "Uncategorized",
            f"{float(txn.amount):.2f}",
            "Income" if float(txn.amount) < 0 else "Expense",
            "Yes" if txn.is_excluded else "No",
            "Yes" if txn.is_transfer else "No",
            txn.notes or ""
        ])

    text_output.flush()
    text_output.detach()
    output.seek(0)

    # Audit log
    audit.log_audit_event(
        db, audit.DATA_EXPORT, user_id=current_user.id,
        details={"format": "csv", "rows": len(transactions)},
    )

    filename = f"transactions_{date.today().isoformat()}.csv"
    return StreamingResponse(
        output,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/transactions/excel")
@limiter.limit("10/minute")
async def export_transactions_excel(
    request: Request,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user=Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Export transactions as Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    profile = get_user_profile(db, current_user)
    account_ids = [a.id for a in db.query(Account).filter(Account.profile_id == profile.id).all()]

    query = db.query(Transaction).options(
        joinedload(Transaction.account),
        joinedload(Transaction.category)
    ).filter(Transaction.account_id.in_(account_ids))

    if start_date:
        query = query.filter(Transaction.date >= start_date)
    if end_date:
        query = query.filter(Transaction.date <= end_date)

    transactions = query.order_by(Transaction.date.desc()).limit(50000).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        bottom=Side(style='thin', color='E5E7EB')
    )
    income_font = Font(color="16A34A")
    expense_font = Font(color="1F2937")

    # Headers
    headers = ["Date", "Description", "Account", "Category", "Amount", "Type", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Data
    for row_idx, txn in enumerate(transactions, 2):
        amount = float(txn.amount)
        ws.cell(row=row_idx, column=1, value=str(txn.date))
        ws.cell(row=row_idx, column=2, value=txn.custom_name or txn.merchant_name or txn.name)
        ws.cell(row=row_idx, column=3, value=(txn.account.display_name or txn.account.name) if txn.account else "")
        ws.cell(row=row_idx, column=4, value=txn.category.name if txn.category else "Uncategorized")

        amount_cell = ws.cell(row=row_idx, column=5, value=amount)
        amount_cell.number_format = '#,##0.00'
        amount_cell.font = income_font if amount < 0 else expense_font

        ws.cell(row=row_idx, column=6, value="Income" if amount < 0 else "Expense")
        ws.cell(row=row_idx, column=7, value=txn.notes or "")

        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).border = thin_border

    # Auto-width columns
    for col in range(1, 8):
        max_length = max(
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(1, min(len(transactions) + 2, 100))
        ) if transactions else 10
        ws.column_dimensions[chr(64 + col)].width = min(max_length + 4, 40)

    # Audit log
    audit.log_audit_event(
        db, audit.DATA_EXPORT, user_id=current_user.id,
        details={"format": "excel", "rows": len(transactions)},
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"transactions_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/report/pdf")
@limiter.limit("5/minute")
async def export_report_pdf(
    request: Request,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user=Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Export a financial summary report as PDF."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from sqlalchemy import func, case

    profile = get_user_profile(db, current_user)
    profile_ids = [p.id for p in current_user.profiles]

    # Date range
    if start_date:
        sd = date.fromisoformat(start_date)
    else:
        today = date.today()
        sd = date(today.year, today.month, 1)
    if end_date:
        ed = date.fromisoformat(end_date)
    else:
        ed = date.today()

    account_ids = [a.id for a in db.query(Account).filter(Account.profile_id.in_(profile_ids)).all()]

    # Summary stats
    income_total = db.query(func.sum(Transaction.amount)).filter(
        Transaction.account_id.in_(account_ids),
        Transaction.date >= sd, Transaction.date <= ed,
        Transaction.is_excluded == False, Transaction.is_transfer == False,
        Transaction.amount < 0
    ).scalar() or 0
    income_total = abs(float(income_total))

    expense_total = db.query(func.sum(Transaction.amount)).filter(
        Transaction.account_id.in_(account_ids),
        Transaction.date >= sd, Transaction.date <= ed,
        Transaction.is_excluded == False, Transaction.is_transfer == False,
        Transaction.amount > 0
    ).scalar() or 0
    expense_total = float(expense_total)

    # Spending by category
    cat_data = db.query(
        Category.name, func.sum(Transaction.amount).label("total")
    ).select_from(Transaction).outerjoin(Category).filter(
        Transaction.account_id.in_(account_ids),
        Transaction.date >= sd, Transaction.date <= ed,
        Transaction.is_excluded == False, Transaction.is_transfer == False,
        Transaction.amount > 0
    ).group_by(Category.name).order_by(func.sum(Transaction.amount).desc()).limit(10).all()

    # Net worth
    latest_nw = db.query(NetWorthSnapshot).filter(
        NetWorthSnapshot.profile_id.in_(profile_ids + [None])
    ).order_by(NetWorthSnapshot.date.desc()).first()

    # Build PDF
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter, topMargin=0.5 * inch, bottomMargin=0.5 * inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("CustomTitle", parent=styles["Title"], fontSize=18, spaceAfter=6)
    subtitle_style = ParagraphStyle("Subtitle", parent=styles["Normal"], fontSize=10, textColor=colors.grey)

    elements = []

    # Title
    elements.append(Paragraph("Financial Report", title_style))
    elements.append(Paragraph(f"{sd.isoformat()} to {ed.isoformat()}", subtitle_style))
    elements.append(Spacer(1, 0.3 * inch))

    # Summary
    elements.append(Paragraph("Summary", styles["Heading2"]))
    summary_data = [
        ["Metric", "Amount"],
        ["Total Income", f"${income_total:,.2f}"],
        ["Total Expenses", f"${expense_total:,.2f}"],
        ["Net Savings", f"${income_total - expense_total:,.2f}"],
    ]
    if latest_nw:
        summary_data.append(["Net Worth", f"${float(latest_nw.net_worth):,.2f}"])

    summary_table = Table(summary_data, colWidths=[3 * inch, 2.5 * inch])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d9488")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3 * inch))

    # Spending by category
    if cat_data:
        elements.append(Paragraph("Top Spending Categories", styles["Heading2"]))
        cat_rows = [["Category", "Amount"]]
        for row in cat_data:
            cat_rows.append([row.name or "Uncategorized", f"${float(row.total):,.2f}"])

        cat_table = Table(cat_rows, colWidths=[3 * inch, 2.5 * inch])
        cat_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d9488")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(cat_table)

    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph(
        f"Generated on {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} by Rundskue Finance Tracker",
        subtitle_style
    ))

    doc.build(elements)
    output.seek(0)

    # Audit log
    audit.log_audit_event(
        db, audit.DATA_EXPORT, user_id=current_user.id,
        details={"format": "pdf", "start_date": str(sd), "end_date": str(ed)},
    )

    filename = f"financial_report_{date.today().isoformat()}.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
